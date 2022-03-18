import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    Sheet = wb['Sheet1']
    cell = Sheet['a1']
    cell = Sheet.cell(1,1)

    for row in range(2, Sheet.max_row + 1):
        cell = Sheet.cell(row, 3)
        corrected_price = cell.value * 0.5
        corrected_price_cell = Sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(Sheet,
                       min_row=2,
                       max_row=Sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    Sheet.add_chart(chart, 'e2')

    wb.save(filename)